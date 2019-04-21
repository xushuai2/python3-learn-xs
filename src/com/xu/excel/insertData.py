#coding=utf-8
'''
Created on 2019年4月20日
python处理excel已经有大量包，主流代表有：
•xlwings：简单强大，可替代VBA

•openpyxl：简单易用，功能广泛

•pandas：使用需要结合其他库，数据处理是pandas立身之本

•win32com：不仅仅是excel，可以处理office;不过它相当于是 windows COM 的封装，新手使用起来略有些痛苦。

•Xlsxwriter：丰富多样的特性，缺点是不能打开/修改已有文件，意味着使用 xlsxwriter 需要从零开始。

•DataNitro：作为插件内嵌到excel中，可替代VBA，在excel中优雅的使用python

•xlutils：结合xlrd/xlwt，老牌python包，需要注意的是你必须同时安装这三个库
@author: 大懒和小懒
'''
import openpyxl
import openpyxl.styles
import random
from openpyxl.utils import get_column_letter
# 打开已有的excel

def createSheet():
    # ------------------ 表操作 ------------------
    # 新建sheet表
    wb.create_sheet(index=2, title="sheet2")  # 可通过index控制创建的表的位置
    wb.create_sheet(index=3, title="sheet3")  # 可通过index控制创建的表的位置
    wb.create_sheet(index=4, title="sheet4")  # 可通过index控制创建的表的位置
    # 获取所有表名
    sheet_names = wb.sheetnames  # 得到工作簿的所有工作表名 结果： ['Sheet1', 'Sheet2', 'Sheet3']
    print(sheet_names)
    # 根据表名删除sheet表
    wb.remove(wb[sheet_names[2]])  # 删除上一步建的第3个工作表
    sheet_names = wb.sheetnames  # 得到工作簿的所有工作表名 结果： ['Sheet1', 'Sheet2', 'Sheet3']
    print(sheet_names)
    


def insert():
    # ------------------ 写入数据 ------------------
    sheet1.cell(row=1, column=2, value="B1")  # 修改第一行第二列的单元格的值为B1
    # sheet1["A1"] = "A1"  # 直接修改A1单元格的值为A1
    sheet1.cell(row = 4, column = 2).value = 'test'
    
    sheet1["B11"] = "B11"  # 新增B11单元格的值为B11
    sheet1.title = "test_sheet"  # 修改sheet1的表名为test_sheet
    print("--- insert数据 ------------------")
    for row in range(1, 20):
        for col in range(1, 54):
            sheet1.cell(column=col, row=row, value=random.choice('abcdefghijklmnopqrstuvwxyz!@#$%^&*()') )




def read():
    # ------------------ 读取数据 ------------------
    print("--- 读取数据 ------------------")
    # 获取单元格数据
    sheet1_max_colum = sheet1.max_column  # 获取最大列数 结果：3
    # ws = wb.active  # 获取当前活动的sheet页
    sheet1_max_row = sheet1.max_row  # 获取最大行数 结果：10
    A1_value = sheet1['A1'].value  # 获取单元格A1值 结果：a1
    A1_column = sheet1['A1'].column  # 获取单元格A1列值 结果： A
    A1_row = sheet1['A1'].row   # 获取单元格A1行号 结果： 1
    A1 = sheet1.cell(row=1, column=1).value  # 获取第一行第一列的单元格值 结果：a1
     
    # 获取C列的所有数据
    list_sheet1_column_C = []
    for i in sheet1["C"]:
        list_sheet1_column_C.append(i.value)
        
    print(list_sheet1_column_C)
     
    # 获取第1行的所有数据
    list_sheet1_row_1 = []
    
    for i in sheet1[1]:
        list_sheet1_row_1.append(i.value)
    print(list_sheet1_row_1) 
    
    # 读取所有数据
    list_sheet1_all = []
    for row in sheet1.rows:
        for cell in row:
            list_sheet1_all.append(cell.value)  # 按行循环获取所有的值，保存在 list_sheet1_all 列表中
    
    print(list_sheet1_all) 


def createWorkBook():
    # 新建一个空excel，表名为sheet，文件名为test
    wb = openpyxl.Workbook()  # 创建新的excel文件，一个工作簿(workbook)在创建的时候同时至少也新建了一张工作表(worksheet)
    wb.save(excelpath)
    print('--------------WorkBook-----creatsed')

def saveFile():        
    # ----------------保存文件
    wb.save(excelpath)

def style():
    # ------------------ 表格样式调整 ------------------
    # 表格样式支持：字体、颜色、模式、边框、数字格式等
     
    # A1单元格 等线24号加粗斜体，字体颜色浅蓝色
    sheet1["B11"].font = openpyxl.styles.Font(name="宋体", size=24, italic=True, color="00CCFF", bold=True)
     
    # B1单元格 水平上下居中
    sheet1["B11"].alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
     
    # 第一行高度设置为30
    sheet1.row_dimensions[1].height = 30
     
    # C列的宽度设置为35
    sheet1.column_dimensions["C"].width = 35



# -------------------------------------------------------------------------------------
excelpath = "E:\\pythonWorkspace\\test.xlsx"

wb = openpyxl.load_workbook(excelpath)

sheet_names = wb.sheetnames
# 根据表名打开sheet表
sheet1 = wb[sheet_names[0]]  # 打开第一个 sheet 工作表

style()
saveFile()



